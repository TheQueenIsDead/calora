"""
Builds the pre-populated Calora SQLite database from NZ/AU food datasets.

Sources:
  - AUSNUT 2023 (Food Standards Australia New Zealand)
  - NZ Food Composition Tables 14th edition (Plant & Food Research)
  - Open Food Facts NZ filtered dump (barcodes + branded products)

Output: ../assets/calora_seed.db  (copied to app on first launch)

Run from the data/ directory:
    uv run python build_db.py
"""

import csv
import json
import re
import sqlite3
import sys
import time
import urllib.request
import uuid
from pathlib import Path

import duckdb
import openpyxl

OFF_API      = "https://world.openfoodfacts.org/api/v2/product"
_API_DELAY   = 0.1   # seconds between API calls (reduce after initial enrichment run)

ROOT = Path(__file__).parent
OUT = ROOT.parent / "assets" / "calora_seed.db"

KJ_TO_KCAL = 1 / 4.184


def _f(v) -> float:
    """Coerce a cell value to float, returning 0 on failure."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


MIGRATIONS_DIR = ROOT / "migrations"


def _apply_migrations(conn: sqlite3.Connection) -> None:
    """Run any migration .sql files not yet recorded in schema_migrations."""
    # Ensure the tracking table exists before anything else.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS schema_migrations (
            version    TEXT PRIMARY KEY,
            applied_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.commit()

    applied = {r[0] for r in conn.execute("SELECT version FROM schema_migrations")}
    for path in sorted(MIGRATIONS_DIR.glob("*.sql")):
        version = path.stem
        if version in applied:
            continue
        print(f"  Applying migration {version} …", flush=True)
        conn.executescript(path.read_text())
        conn.execute("INSERT INTO schema_migrations (version) VALUES (?)", (version,))
        conn.commit()


def _insert(conn: sqlite3.Connection, row: dict, replace: bool = False) -> None:
    verb = "INSERT OR REPLACE" if replace else "INSERT OR IGNORE"
    conn.execute(
        f"""
        {verb} INTO foods
            (id, name, brand, barcode, calories_per_100g,
             fat_per_100g, saturated_fat_per_100g, carbs_per_100g,
             sugars_per_100g, fiber_per_100g, protein_per_100g,
             sodium_per_100g, serving_size, serving_grams, image_url, source)
        VALUES
            (:id, :name, :brand, :barcode, :calories_per_100g,
             :fat_per_100g, :saturated_fat_per_100g, :carbs_per_100g,
             :sugars_per_100g, :fiber_per_100g, :protein_per_100g,
             :sodium_per_100g, :serving_size, :serving_grams, :image_url, :source)
        """,
        row,
    )


# ---------------------------------------------------------------------------
# AUSNUT 2023
# ---------------------------------------------------------------------------

def import_ausnut(conn: sqlite3.Connection) -> int:
    base = ROOT / "ausnut"

    # --- food details: Survey ID → food name ---
    wb_details = openpyxl.load_workbook(base / "AUSNUT 2023 - Food details.xlsx", read_only=True)
    ws_details = wb_details["Food details"]
    names: dict[int, str] = {}
    for i, row in enumerate(ws_details.iter_rows(values_only=True)):
        if i < 3:  # skip title rows + header row (0-indexed row 2 = header)
            continue
        survey_id = row[0]
        food_name = row[4]
        if survey_id and food_name:
            names[int(survey_id)] = str(food_name)
    wb_details.close()

    # --- nutrient profiles: one row per food ---
    wb_nutr = openpyxl.load_workbook(base / "AUSNUT 2023 - Food nutrient profiles.xlsx", read_only=True)
    ws_nutr = wb_nutr["Food nutrient profiles"]

    count = 0
    for i, row in enumerate(ws_nutr.iter_rows(values_only=True)):
        if i < 3:
            continue
        survey_id = row[0]
        if not survey_id:
            continue
        sid = int(survey_id)
        name = str(row[3]) if row[3] else names.get(sid, "")
        if not name:
            continue

        energy_kj = _f(row[4])
        if energy_kj <= 0:
            continue

        _insert(conn, {
            "id": f"ausnut_{sid}",
            "name": name,
            "brand": None,
            "barcode": None,
            "calories_per_100g": round(energy_kj * KJ_TO_KCAL, 2),
            "protein_per_100g": _f(row[7]),
            "fat_per_100g": _f(row[8]),
            "saturated_fat_per_100g": _f(row[49]),
            "carbs_per_100g": _f(row[9]),
            "sugars_per_100g": _f(row[12]),
            "fiber_per_100g": _f(row[15]),
            "sodium_per_100g": round(_f(row[25]) / 1000, 4),  # mg → g
            "serving_size": None,
            "serving_grams": None,
            "image_url": None,
            "source": "ausnut",
        }, replace=True)
        count += 1

    wb_nutr.close()
    return count


# ---------------------------------------------------------------------------
# NZ Food Composition Tables 14th Edition
# ---------------------------------------------------------------------------

def import_nz_composition(conn: sqlite3.Connection) -> int:
    path = ROOT / "nzfc" / "concise-14-edition.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    count = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:  # rows 0-4: header, blank, units, blank, first category
            continue
        food_id = row[0]
        # Skip None rows (per-serving alternates) and single-letter category headers
        if not food_id:
            continue
        food_id_str = str(food_id).strip()
        if len(food_id_str) <= 1 or not any(c.isdigit() for c in food_id_str):
            continue

        name = str(row[1]).strip() if row[1] else ""
        if not name:
            continue

        energy_kj = _f(row[4])
        if energy_kj <= 0:
            continue

        _insert(conn, {
            "id": f"nz_{food_id_str}",
            "name": name,
            "brand": None,
            "barcode": None,
            "calories_per_100g": round(energy_kj * KJ_TO_KCAL, 2),
            "protein_per_100g": _f(row[6]),
            "fat_per_100g": _f(row[7]),
            "saturated_fat_per_100g": _f(row[12]),
            "carbs_per_100g": _f(row[8]),
            "sugars_per_100g": _f(row[10]),
            "fiber_per_100g": _f(row[9]),
            "sodium_per_100g": round(_f(row[18]) / 1000, 4),  # mg → g
            "serving_size": None,
            "serving_grams": _f(row[2]) or None,
            "image_url": None,
            "source": "nz",
        }, replace=True)
        count += 1

    wb.close()
    return count


# ---------------------------------------------------------------------------
# Open Food Facts APAC — CSV with API enrichment for missing nutrition
# ---------------------------------------------------------------------------


def _fetch_off_nutrition(barcode: str) -> tuple | None:
    """Fetch a single product from the OFF API. Returns a nutrition tuple or None."""
    try:
        url = f"{OFF_API}/{barcode}?fields=product_name,brands,serving_size,serving_quantity,nutriments,image_front_url"
        req = urllib.request.Request(
            url, headers={"User-Agent": "Calora/1.0 (calorie tracker; data pipeline)"}
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        if data.get("status") != 1:
            return None
        p = data["product"]
        n = p.get("nutriments") or {}
        kcal = n.get("energy-kcal_100g") or 0
        if not kcal:
            kj = n.get("energy_100g") or 0
            kcal = round(kj / 4.184, 2) if kj else 0
        if kcal <= 0:
            return None
        try:
            serving_grams = float(p.get("serving_quantity") or "") or None
        except ValueError:
            serving_grams = None
        return (
            (p.get("product_name") or "").strip(),
            (p.get("brands") or "").strip() or None,
            round(kcal, 2),
            n.get("fat_100g") or 0,
            n.get("saturated-fat_100g") or 0,
            n.get("carbohydrates_100g") or 0,
            n.get("sugars_100g") or 0,
            n.get("fiber_100g") or 0,
            n.get("proteins_100g") or 0,
            n.get("sodium_100g") or 0,
            (p.get("serving_size") or "").strip() or None,
            serving_grams,
            (p.get("image_front_url") or "").strip() or None,
        )
    except Exception:
        return None


def _insert_off(conn: sqlite3.Connection, barcode: str, name: str,
                brand, kcal: float, fat: float, sat_fat: float,
                carbs: float, sugars: float, fiber: float, protein: float,
                sodium: float, serving_size, serving_grams, image_url) -> None:
    _insert(conn, {
        "id":                     f"off_{barcode}",
        "name":                   name,
        "brand":                  brand,
        "barcode":                barcode,
        "calories_per_100g":      round(kcal, 2),
        "fat_per_100g":           fat,
        "saturated_fat_per_100g": sat_fat,
        "carbs_per_100g":         carbs,
        "sugars_per_100g":        sugars,
        "fiber_per_100g":         fiber,
        "protein_per_100g":       protein,
        "sodium_per_100g":        sodium,
        "serving_size":           serving_size,
        "serving_grams":          serving_grams,
        "image_url":              image_url,
        "source":                 "off_nz",
    })


def import_off_parquet(conn: sqlite3.Connection) -> int:
    parquet_path = ROOT / "off" / "food.parquet"
    if not parquet_path.exists():
        raise FileNotFoundError(f"{parquet_path}\n  Run: task parquet")

    print(f"  Querying {parquet_path.name} for NZ/AU products …", end=" ", flush=True)
    con = duckdb.connect()
    rows = con.execute(f"""
        SELECT
            code,
            product_name[1].text                                                AS name,
            brands,
            serving_size,
            serving_quantity,
            coalesce(
                nullif([x for x in nutriments if x.name = 'energy-kcal'][1]['100g'], 0),
                [x for x in nutriments if x.name = 'energy-kj'][1]['100g'] / 4.184
            )                                                                   AS kcal,
            [x for x in nutriments if x.name = 'fat'][1]['100g']               AS fat,
            [x for x in nutriments if x.name = 'saturated-fat'][1]['100g']     AS sat_fat,
            [x for x in nutriments if x.name = 'carbohydrates'][1]['100g']     AS carbs,
            [x for x in nutriments if x.name = 'sugars'][1]['100g']            AS sugars,
            [x for x in nutriments if x.name = 'fiber'][1]['100g']             AS fiber,
            [x for x in nutriments if x.name = 'proteins'][1]['100g']          AS protein,
            [x for x in nutriments if x.name = 'sodium'][1]['100g']            AS sodium
        FROM '{parquet_path}'
        WHERE (list_contains(countries_tags, 'en:new-zealand')
            OR list_contains(countries_tags, 'en:australia'))
          AND code IS NOT NULL
          AND len(product_name) > 0
          AND NOT list_contains(misc_tags, 'en:nutriscore-missing-nutrition-data-energy')
    """).fetchall()
    con.close()
    print(f"{len(rows):,} products", flush=True)

    count        = 0
    no_nutrition = 0
    total_rows   = len(rows)
    LOG_EVERY    = 5_000
    COMMIT_EVERY = 500

    for i, (code, name, brands, serving_size, serving_qty,
            kcal, fat, sat_fat, carbs, sugars, fiber, protein, sodium) in enumerate(rows, 1):

        name = (name or "").strip()
        if not code or not name:
            continue

        if kcal is not None:
            try:
                serving_grams = float(serving_qty) if serving_qty else None
            except (TypeError, ValueError):
                serving_grams = None
            _insert_off(
                conn, code, name,
                (brands or "").strip() or None,
                round(kcal, 2),
                fat or 0, sat_fat or 0, carbs or 0,
                sugars or 0, fiber or 0, protein or 0, sodium or 0,
                (serving_size or "").strip() or None,
                serving_grams, None,
            )
            count += 1

        else:
            existing = conn.execute(
                "SELECT 1 FROM foods WHERE id = ? AND calories_per_100g >= 0",
                (f"off_{code}",)
            ).fetchone()
            if existing:
                count += 1  # enriched in a previous build — keep it
            else:
                no_nutrition += 1

        if i % COMMIT_EVERY == 0:
            conn.commit()

        if i % LOG_EVERY == 0:
            total_db = conn.execute("SELECT COUNT(*) FROM foods WHERE source='off_nz'").fetchone()[0]
            print(
                f"  {i:>6,} / {total_rows:,}  |  {total_db:,} in DB"
                f"  |  {no_nutrition:,} missing nutrition",
                flush=True,
            )

    conn.commit()
    print(
        f"  Done — {total_rows:,} rows  |  {count:,} inserted"
        f"  |  {no_nutrition:,} skipped (no nutrition in parquet)",
        flush=True,
    )
    return count


# ---------------------------------------------------------------------------
# USDA Foundation Foods 2026
# ---------------------------------------------------------------------------

def import_usda(conn: sqlite3.Connection) -> int:
    base = ROOT / "usda"
    if not (base / "food.csv").exists():
        raise FileNotFoundError(f"{base / 'food.csv'}\n  Run: task usda")

    # nutrient_id → field name (prefer 1008 "Energy KCAL", fall back to Atwater)
    ENERGY_IDS   = {'1008', '2047', '2048'}
    NUTRIENT_MAP = {
        '1003': 'protein_per_100g',
        '1004': 'fat_per_100g',
        '1005': 'carbs_per_100g',
        '1079': 'fiber_per_100g',
        '1093': 'sodium_mg',       # mg — convert later
        '1258': 'saturated_fat_per_100g',
        '2000': 'sugars_per_100g',
    }

    # fdc_id → description
    foods: dict[str, str] = {}
    with open(base / "food.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            foods[row['fdc_id']] = row['description']

    # fdc_id → {field: value}
    nutrients: dict[str, dict] = {fid: {} for fid in foods}
    with open(base / "food_nutrient.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            fid = row['fdc_id']
            nid = row['nutrient_id']
            amt = _f(row['amount'])
            if fid not in nutrients:
                continue
            if nid in ENERGY_IDS and 'calories_per_100g' not in nutrients[fid]:
                nutrients[fid]['calories_per_100g'] = amt
            elif nid in NUTRIENT_MAP:
                field = NUTRIENT_MAP[nid]
                if field not in nutrients[fid]:
                    nutrients[fid][field] = amt

    count = 0
    for fid, name in foods.items():
        n = nutrients.get(fid, {})
        kcal = n.get('calories_per_100g', 0)
        if not name or kcal <= 0:
            continue
        sodium_g = round(n.get('sodium_mg', 0) / 1000, 4)
        _insert(conn, {
            'id': f'usda_{fid}',
            'name': name,
            'brand': None,
            'barcode': None,
            'calories_per_100g': round(kcal, 2),
            'protein_per_100g': n.get('protein_per_100g', 0),
            'fat_per_100g': n.get('fat_per_100g', 0),
            'saturated_fat_per_100g': n.get('saturated_fat_per_100g', 0),
            'carbs_per_100g': n.get('carbs_per_100g', 0),
            'sugars_per_100g': n.get('sugars_per_100g', 0),
            'fiber_per_100g': n.get('fiber_per_100g', 0),
            'sodium_per_100g': sodium_g,
            'serving_size': None,
            'serving_grams': None,
            'image_url': None,
            'source': 'usda',
        }, replace=True)
        count += 1
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def _counts(db_path: Path) -> dict[str, int]:
    """Return per-source row counts from an existing DB, or empty dict."""
    if not db_path.exists():
        return {}
    try:
        c = sqlite3.connect(db_path)
        rows = c.execute("SELECT source, COUNT(*) FROM foods GROUP BY source").fetchall()
        c.close()
        return {source: count for source, count in rows}
    except Exception:
        return {}


def _print_diff(before: dict[str, int], after: dict[str, int]) -> None:
    sources = sorted(set(before) | set(after))
    total_before = sum(before.values())
    total_after  = sum(after.values())
    print("\n── Record counts ─────────────────────────────")
    print(f"  {'Source':<20} {'Before':>8}  {'After':>8}  {'Δ':>8}")
    print(f"  {'─'*20}  {'─'*8}  {'─'*8}  {'─'*8}")
    for src in sources:
        b, a = before.get(src, 0), after.get(src, 0)
        delta = a - b
        flag  = "  +" if delta > 0 else ("  -" if delta < 0 else "")
        print(f"  {src:<20} {b:>8,}  {a:>8,}  {delta:>+8,}{flag}")
    print(f"  {'─'*20}  {'─'*8}  {'─'*8}  {'─'*8}")
    print(f"  {'TOTAL':<20} {total_before:>8,}  {total_after:>8,}  {total_after - total_before:>+8,}")
    print("──────────────────────────────────────────────")


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    before = _counts(OUT)

    conn = sqlite3.connect(OUT)
    conn.execute("PRAGMA journal_mode=WAL")
    print("Applying migrations …")
    _apply_migrations(conn)

    print("Importing AUSNUT 2023 …", end=" ", flush=True)
    n = import_ausnut(conn)
    conn.commit()
    print(f"{n:,} rows")

    print("Importing NZ Food Composition …", end=" ", flush=True)
    n = import_nz_composition(conn)
    conn.commit()
    print(f"{n:,} rows")

    print("Importing Open Food Facts APAC (Parquet) …", flush=True)
    n = import_off_parquet(conn)
    conn.commit()
    print(f"  {n:,} rows total")

    print("Importing USDA Foundation Foods …", end=" ", flush=True)
    n = import_usda(conn)
    conn.commit()
    print(f"{n:,} rows")

    after = dict(conn.execute("SELECT source, COUNT(*) FROM foods GROUP BY source").fetchall())
    _print_diff(before, after)

    print("Rebuilding vocabulary …", end=" ", flush=True)
    conn.execute("DROP TABLE IF EXISTS vocabulary")
    tokens: set[str] = set()
    for (name, brand) in conn.execute("SELECT name, COALESCE(brand,'') FROM foods"):
        for word in re.findall(r'[a-z]{3,}', f"{name} {brand}".lower()):
            tokens.add(word)
    conn.execute("CREATE TABLE vocabulary (word TEXT PRIMARY KEY)")
    conn.executemany("INSERT OR IGNORE INTO vocabulary (word) VALUES (?)", [(w,) for w in tokens])
    conn.commit()
    print(f"{len(tokens):,} tokens")

    print("Rebuilding FTS5 index …", end=" ", flush=True)
    conn.execute("DROP TABLE IF EXISTS foods_fts")
    conn.execute("""
        CREATE VIRTUAL TABLE foods_fts USING fts5(
            name,
            brand,
            content=foods,
            content_rowid=rowid,
            tokenize='unicode61 remove_diacritics 2'
        )
    """)
    conn.execute("INSERT INTO foods_fts(foods_fts) VALUES('rebuild')")
    conn.commit()
    print("done")

    print(f"Output: {OUT}")
    conn.execute("PRAGMA optimize")
    conn.execute("PRAGMA user_version = 4")  # bump when schema or data changes
    conn.execute("VACUUM")
    conn.close()


if __name__ == "__main__":
    main()